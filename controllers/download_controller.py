
from flask import send_file
import shutil
import openpyxl
from openpyxl import load_workbook

class DownloadController:

    def __init__(self,parameterService,tickerRateService):
        self.parameterService = parameterService
        self.tickerRateService = tickerRateService
        pass

    def dispatch(self, request):
        tickers, from_date, till_date = self.parameterService.init_params()
        tickers = "aapl"
        ticker_data = self.tickerRateService.get_rates(tickers, from_date, till_date)

        dest_file = 'static/excel/excel_dummy2.xlsm'

        shutil.copy('static/excel/excel_dummy1.xlsm', dest_file)

        wb = load_workbook(filename=dest_file)
        ws = wb["Summary"]
        ws["b4"] = tickers
        ws["b5"] = str(from_date)
        ws["b6"] = str(till_date)
        wb.save(dest_file)

        return send_file(dest_file,
                         mimetype='text/csv',
                         attachment_filename='dummy.xlsm',
                         as_attachment=True)




